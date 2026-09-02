"""
Coletor de leads — Google Maps (Places API) v3
------------------------------------------------------------------------
Novidades desta versão:
  1. Pula estabelecimentos já coletados em execuções anteriores (usa um
     arquivo de cache local com os IDs já processados — economiza API
     também, porque nem chama os Detalhes de quem já foi visto).
  2. Adiciona as colunas "possui_site" (Sim/Não) e "site" (link, se tiver).
     Isso já vem de graça na resposta da Places API, não precisa de busca
     extra.
  3. Exporta pra Excel (.xlsx) ou direto pro Google Sheets — escolha em
     MODO_SAIDA.
  4. Configuração separada: variáveis não sensíveis ficam em config.py e
     variáveis sensíveis (API key, e-mails, IDs) ficam no arquivo .env.

CONFIGURAÇÃO NECESSÁRIA
  pip install requests openpyxl

  Se for usar Google Sheets (MODO_SAIDA = "sheets"), instale também:
  pip install gspread google-auth

  Configuração (config.py + .env):
    - config.py: parâmetros não sensíveis (buscas, delays, colunas, etc.)
    - .env: API_KEY e demais credenciais/identificadores.

  Places API (New):
  https://console.cloud.google.com/apis/library/places-backend.googleapis.com
  -> gera uma API_KEY e coloque no arquivo .env

  Google Sheets (só se for usar esse modo):
    1. Ative as APIs (mesmo projeto do Google Cloud):
       https://console.cloud.google.com/apis/library/sheets.googleapis.com
       https://console.cloud.google.com/apis/library/drive.googleapis.com
    2. Crie uma Service Account em "IAM e administrador > Contas de
       serviço" > Criar. Depois, na aba "Chaves" dela, gere uma chave
       JSON e baixe o arquivo.
    3. Aponte SERVICE_ACCOUNT_FILE pro caminho desse .json (no .env).
    4. Preencha MEU_EMAIL com seu e-mail (no .env) — sem isso a planilha
       fica visível só pra service account, e você não consegue abrir.
    5. Deixe PLANILHA_ID em branco na primeira execução: o script cria a
       planilha, compartilha com você e imprime o ID no console. Copie
       esse ID pra PLANILHA_ID no .env nas próximas execuções (senão ele
       cria uma planilha nova toda vez, em vez de continuar a mesma).

AVISOS IMPORTANTES
  - Cota: a Places API tem cota gratuita mensal e cobra por chamada
    depois disso — considere limitar a cota diária no Google Cloud
    Console (Cotas) pra nunca ultrapassar o quanto você quer gastar.
  - Ao visitar os sites das empresas (enriquecimento), o script respeita
    o robots.txt de cada site e usa um delay entre requisições.
  - LGPD: mesmo em prospecção B2B, guarde a base legal do tratamento
    (legítimo interesse costuma ser a mais usada aqui), ofereça opção de
    descadastro no primeiro contato, e para ligações/SMS consulte a
    lista "Não Me Perturbe" (naomeperturbe.com.br) antes de discar.
"""

import json
import os
import re
import sys
import time
import requests
from urllib.parse import urlparse
from urllib.robotparser import RobotFileParser

from config import (
    API_KEY,
    ARQUIVO_EXCEL,
    CAMPOS,
    DELAY_ENTRE_REQUISICOES,
    IDS_CACHE_ARQUIVO,
    MAPS_BUSCAS,
    MEU_EMAIL,
    MODO_SAIDA,
    PLANILHA_ID,
    PLANILHA_NOME,
    RESULTADOS_POR_BUSCA,
    SERVICE_ACCOUNT_FILE,
    TIMEOUT_SITE,
)

# ----------------------------------------------------------------------
# CACHE DE IDS JÁ COLETADOS (feature 1)
# ----------------------------------------------------------------------

print(f"Resultados por busca: {RESULTADOS_POR_BUSCA}")

def carregar_ids_vistos():
    if os.path.exists(IDS_CACHE_ARQUIVO):
        with open(IDS_CACHE_ARQUIVO, "r", encoding="utf-8") as f:
            return set(json.load(f))
    return set()


def salvar_ids_vistos(ids_vistos):
    with open(IDS_CACHE_ARQUIVO, "w", encoding="utf-8") as f:
        json.dump(sorted(ids_vistos), f)


# ----------------------------------------------------------------------
# GOOGLE PLACES API (Maps)
# ----------------------------------------------------------------------

PLACES_SEARCH_URL = "https://places.googleapis.com/v1/places:searchText"
PLACES_DETAILS_URL = "https://places.googleapis.com/v1/places/{place_id}"

HEADERS_PLACES_SEARCH = {
    "Content-Type": "application/json",
    "X-Goog-Api-Key": API_KEY,
    "X-Goog-FieldMask": "places.id,places.displayName",
}

HEADERS_PLACES_DETAILS = {
    "X-Goog-Api-Key": API_KEY,
    "X-Goog-FieldMask": "displayName,nationalPhoneNumber,websiteUri,googleMapsUri",
}


def buscar_no_maps(query, max_resultados=RESULTADOS_POR_BUSCA):
    body = {"textQuery": query, "languageCode": "pt-BR", "maxResultCount": max_resultados}
    resp = requests.post(PLACES_SEARCH_URL, json=body, headers=HEADERS_PLACES_SEARCH, timeout=15)
    resp.raise_for_status()
    print(f"Resultados da busca: {resp.__str__()}")
    return resp.json().get("places", [])


def detalhes_do_lugar(place_id):
    url = PLACES_DETAILS_URL.format(place_id=place_id)
    resp = requests.get(url, headers=HEADERS_PLACES_DETAILS, timeout=15)
    resp.raise_for_status()
    print(f"Detalhes do lugar: {resp.__str__()}")
    return resp.json()


# ----------------------------------------------------------------------
# ENRIQUECIMENTO: visita o site da empresa e tenta achar contato
# ----------------------------------------------------------------------

REGEX_EMAIL = re.compile(r"[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}")
REGEX_WHATSAPP = re.compile(r"(?:wa\.me/|api\.whatsapp\.com/send\?phone=)(\d{10,13})")
REGEX_TELEFONE = re.compile(r"(?:\+?55\s?)?\(?\d{2}\)?[\s.-]?9?\d{4}[\s.-]?\d{4}")
EXTENSOES_IGNORAR = (".png", ".jpg", ".jpeg", ".gif", ".svg", ".webp")


def sanitizar_url(url):
    """Remove query string e fragmento da URL (ex.: ?utm_source=..., #topo)."""
    if not url:
        return url
    try:
        partes = urlparse(url)
        if partes.scheme in ("http", "https"):
            return f"{partes.scheme}://{partes.netloc}{partes.path}"
        return url
    except Exception:
        return url


def pode_acessar(url):
    try:
        partes = urlparse(url)
        robots_url = f"{partes.scheme}://{partes.netloc}/robots.txt"
        rp = RobotFileParser()
        rp.set_url(robots_url)
        rp.read()
        return rp.can_fetch("*", url)
    except Exception:
        return True


def extrair_contatos_do_site(url):
    contato = {"email": "", "whatsapp": "", "telefone_site": ""}
    if not url or not pode_acessar(url):
        return contato

    try:
        resp = requests.get(
            url, timeout=TIMEOUT_SITE,
            headers={"User-Agent": "Mozilla/5.0 (compatible; LeadBot/1.0)"},
        )
        html = resp.text
    except Exception:
        return contato

    email_match = REGEX_EMAIL.search(html)
    if email_match and not email_match.group(0).lower().endswith(EXTENSOES_IGNORAR):
        contato["email"] = email_match.group(0)

    whats_match = REGEX_WHATSAPP.search(html)
    if whats_match:
        contato["whatsapp"] = whats_match.group(1)

    tel_match = REGEX_TELEFONE.search(html)
    if tel_match:
        contato["telefone_site"] = tel_match.group(0)

    return contato


# ----------------------------------------------------------------------
# COLETA PRINCIPAL
# ----------------------------------------------------------------------

def coletar():
    ids_vistos = carregar_ids_vistos()
    leads_novos = []

    for busca in MAPS_BUSCAS:
        print(f"[Maps] Buscando: {busca['query']}")
        for lugar in buscar_no_maps(busca["query"]):
            place_id = lugar.get("id")
            if not place_id:
                continue
            if place_id in ids_vistos:
                continue  # feature 1: já coletado antes, pula sem gastar chamada de Detalhes

            time.sleep(DELAY_ENTRE_REQUISICOES)
            det = detalhes_do_lugar(place_id)

            site = sanitizar_url(det.get("websiteUri", ""))
            possui_site = "Sim" if site else "Não"

            enriquecido = extrair_contatos_do_site(site)
            time.sleep(DELAY_ENTRE_REQUISICOES)

            telefone = enriquecido["whatsapp"] or det.get("nationalPhoneNumber", "") or enriquecido["telefone_site"]

            leads_novos.append({
                "nome": det.get("displayName", {}).get("text", ""),
                "telefone_whatsapp": telefone,
                "email": enriquecido["email"],
                "tipo_negocio": busca["tipo_negocio"],
                "possui_site": possui_site,
                "site": site,
                "link_origem": det.get("googleMapsUri", ""),
                "fonte": "Google Maps",
            })

            ids_vistos.add(place_id)

    salvar_ids_vistos(ids_vistos)
    return leads_novos


# ----------------------------------------------------------------------
# SAÍDA: EXCEL
# ----------------------------------------------------------------------

def salvar_excel(leads, caminho=ARQUIVO_EXCEL):
    if not leads:
        print("Nenhum lead novo pra salvar.")
        return

    import openpyxl
    from openpyxl.styles import Font

    if os.path.exists(caminho):
        wb = openpyxl.load_workbook(caminho)
        ws = wb.active
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Leads"
        ws.append(CAMPOS)
        for cel in ws[1]:
            cel.font = Font(bold=True)
        ws.freeze_panes = "A2"

    for lead in leads:
        ws.append([lead[c] for c in CAMPOS])

    for i, campo in enumerate(CAMPOS, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(14, len(campo) + 4)

    wb.save(caminho)
    print(f"{len(leads)} leads novos adicionados em {caminho}")


# ----------------------------------------------------------------------
# SAÍDA: GOOGLE SHEETS
# ----------------------------------------------------------------------

def salvar_sheets(leads):
    if not leads:
        print("Nenhum lead novo pra salvar.")
        return

    import gspread
    from google.oauth2.service_account import Credentials

    escopos = ["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
    creds = Credentials.from_service_account_file(SERVICE_ACCOUNT_FILE, scopes=escopos)
    gc = gspread.authorize(creds)

    if PLANILHA_ID:
        sh = gc.open_by_key(PLANILHA_ID)
        ws = sh.sheet1
    else:
        sh = gc.create(PLANILHA_NOME)
        sh.share(MEU_EMAIL, perm_type="user", role="writer")
        ws = sh.sheet1
        ws.append_row(CAMPOS)
        print(f"Planilha criada! Copie este ID pra PLANILHA_ID no script: {sh.id}")
        print(f"Link: https://docs.google.com/spreadsheets/d/{sh.id}")

    linhas = [[lead[c] for c in CAMPOS] for lead in leads]
    ws.append_rows(linhas)
    print(f"{len(leads)} leads novos adicionados na planilha.")


# ----------------------------------------------------------------------

if __name__ == "__main__":
    if not API_KEY:
        print("API_KEY não configurada. Preencha o arquivo .env na raiz do projeto.")
        sys.exit(1)

    leads = coletar()
    if MODO_SAIDA == "sheets":
        salvar_sheets(leads)
    else:
        salvar_excel(leads)